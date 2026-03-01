import os
import shutil
from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction
from django.contrib.auth import get_user_model

from shop.models import (
    UserProfile,
    Category,
    Unit,
    Producer,
    Supplier,
    Product,
    PickupPoint,
    Order,
    OrderItem,
)

User = get_user_model()


def safe_str(val):
    if val is None:
        return ''
    return str(val).strip()


def safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def safe_decimal(val, default=0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


class Command(BaseCommand):
    help = 'Загрузка данных из Excel (Tovar, user_import, Заказ, Пункты выдачи)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--dir',
            type=str,
            default=None,
            help='Папка с файлами import (по умолчанию IMPORT_RESOURCES)',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        base_dir = options.get('dir') or getattr(settings, 'IMPORT_RESOURCES', None)
        if not base_dir:
            base_dir = Path(settings.BASE_DIR) / 'Прил_ОЗ_КОД 09.02.07-2-2026' / 'БУ' / 'Модуль 1' / 'import'
        base_dir = Path(base_dir)
        if not base_dir.exists():
            self.stderr.write(self.style.ERROR(f'Папка не найдена: {base_dir}'))
            return

        path_pv = base_dir / 'Пункты выдачи_import.xlsx'
        if path_pv.exists():
            wb = openpyxl.load_workbook(path_pv, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None:
                    addr = safe_str(row[0])
                    if addr:
                        PickupPoint.objects.get_or_create(address=addr)
            wb.close()
            self.stdout.write(self.style.SUCCESS('Пункты выдачи загружены.'))

        path_user = base_dir / 'user_import.xlsx'
        if path_user.exists():
            wb = openpyxl.load_workbook(path_user, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                role_str, full_name, email, password = (
                    safe_str(row[0]),
                    safe_str(row[1]) if len(row) > 1 else '',
                    safe_str(row[2]) if len(row) > 2 else '',
                    safe_str(row[3]) if len(row) > 3 else '',
                )
                if not email:
                    continue
                role_map = {
                    'Администратор': UserProfile.ROLE_ADMIN,
                    'Менеджер': UserProfile.ROLE_MANAGER,
                    'Клиент': UserProfile.ROLE_CLIENT,
                }
                role = role_map.get(role_str, UserProfile.ROLE_CLIENT)
                username = email.replace('@', '_at_')[:150]
                user, created = User.objects.get_or_create(
                    username=username,
                    defaults={'email': email, 'is_staff': role == UserProfile.ROLE_ADMIN},
                )
                if created:
                    user.set_password(password or 'import')
                    user.save()
                UserProfile.objects.get_or_create(
                    user=user,
                    defaults={'role': role, 'full_name': full_name or username},
                )
            wb.close()
            self.stdout.write(self.style.SUCCESS('Пользователи загружены.'))

        static_img = Path(settings.BASE_DIR) / 'static' / 'img'
        static_img.mkdir(parents=True, exist_ok=True)
        picture_src = base_dir / 'picture.png'
        if picture_src.exists():
            shutil.copy(picture_src, static_img / 'picture.png')
        for icon_name in ('Icon.png', 'Icon.jpg', 'Icon.ico'):
            icon_src = base_dir / icon_name
            if icon_src.exists():
                shutil.copy(icon_src, static_img / icon_name)
        media_products = Path(settings.MEDIA_ROOT) / 'products'
        media_products.mkdir(parents=True, exist_ok=True)

        path_tovar = base_dir / 'Tovar.xlsx'
        if path_tovar.exists():
            wb = openpyxl.load_workbook(path_tovar, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                if len(row) < 10:
                    continue
                article = safe_str(row[0])
                name = safe_str(row[1])
                unit_name = safe_str(row[2])
                price = safe_decimal(row[3])
                producer_name = safe_str(row[4])
                supplier_name = safe_str(row[5])
                category_name = safe_str(row[6])
                discount = safe_int(row[7])
                quantity = safe_int(row[8])
                description = safe_str(row[9]) if len(row) > 9 else ''
                photo = safe_str(row[10]) if len(row) > 10 else ''

                if not article:
                    continue
                if not category_name:
                    category_name = 'Без категории'
                if not unit_name:
                    unit_name = 'шт.'

                category, _ = Category.objects.get_or_create(name=category_name)
                unit, _ = Unit.objects.get_or_create(name=unit_name)
                producer, _ = Producer.objects.get_or_create(name=producer_name)
                supplier, _ = Supplier.objects.get_or_create(name=supplier_name)

                image_path = ''
                if photo:
                    src = base_dir / photo
                    if src.exists():
                        dest = media_products / photo
                        shutil.copy(src, dest)
                        image_path = f'products/{photo}'

                Product.objects.update_or_create(
                    article=article,
                    defaults={
                        'name': name or article,
                        'category': category,
                        'producer': producer,
                        'supplier': supplier,
                        'unit': unit,
                        'price': max(0, price),
                        'quantity': max(0, quantity),
                        'discount': max(0, min(100, discount)),
                        'description': description,
                        'image': image_path,
                    },
                )
            wb.close()
            self.stdout.write(self.style.SUCCESS('Товары загружены.'))

        path_order = base_dir / 'Заказ_import.xlsx'
        if path_order.exists():
            wb = openpyxl.load_workbook(path_order, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                order_id = safe_int(row[0])
                composition = safe_str(row[1]) if len(row) > 1 else ''
                date_created = row[2] if len(row) > 2 else None
                date_delivery = row[3] if len(row) > 3 else None
                pickup_index = safe_int(row[4]) if len(row) > 4 else None
                client_name = safe_str(row[5]) if len(row) > 5 else ''
                number = safe_str(row[6]) if len(row) > 6 else str(order_id)
                status_str = safe_str(row[7]) if len(row) > 7 else 'Ожидает'

                if not number:
                    number = f'ORD-{order_id}'

                status_map = {'Ожидает': Order.STATUS_PENDING, 'Доставлен': Order.STATUS_DELIVERED, 'Отменён': Order.STATUS_CANCELLED}
                status = status_map.get(status_str, Order.STATUS_PENDING)

                pickup_point = None
                if pickup_index is not None and pickup_index > 0:
                    points = list(PickupPoint.objects.all())
                    if 1 <= pickup_index <= len(points):
                        pickup_point = points[pickup_index - 1]

                from datetime import datetime
                if hasattr(date_created, 'date'):
                    dc = date_created
                elif isinstance(date_created, str) and date_created:
                    try:
                        dc = datetime.strptime(date_created[:10], '%Y-%m-%d').date()
                    except Exception:
                        dc = datetime.now().date()
                else:
                    dc = datetime.now().date()

                dd = None
                if date_delivery:
                    if hasattr(date_delivery, 'date'):
                        dd = date_delivery
                    elif isinstance(date_delivery, str) and date_delivery:
                        try:
                            dd = datetime.strptime(date_delivery[:10], '%Y-%m-%d').date()
                        except Exception:
                            pass

                order, created = Order.objects.get_or_create(
                    number=number,
                    defaults={
                        'client_name': client_name,
                        'date_created': dc,
                        'date_delivery': dd,
                        'pickup_point': pickup_point,
                        'status': status,
                    },
                )
                if not created or not composition:
                    continue
                parts = [p.strip() for p in composition.split(',')]
                i = 0
                while i + 1 < len(parts):
                    art = parts[i]
                    try:
                        qty = int(parts[i + 1])
                    except (ValueError, IndexError):
                        i += 1
                        continue
                    product = Product.objects.filter(article=art).first()
                    if product and qty > 0:
                        OrderItem.objects.get_or_create(
                            order=order,
                            product=product,
                            defaults={'quantity': qty},
                        )
                    i += 2
            wb.close()
            self.stdout.write(self.style.SUCCESS('Заказы загружены.'))

        self.stdout.write(self.style.SUCCESS('Импорт завершён.'))
